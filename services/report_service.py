"""
CarbonLens V8 — Report service.
Assembles all export formats from a single ComputedState.
No PDF rendering yet (Sprint 5). No formulas. No st.* imports.
"""
from __future__ import annotations
import datetime
import json
import logging
from typing import Optional

log = logging.getLogger("carbonlens.services.report")


def build_report_context(state: dict, org: dict) -> dict:
    """
    Build the unified report context dict from a ComputedState.

    This is the SINGLE data source for ALL export formats.
    No export format may query ComputedState directly.

    Parameters
    ----------
    state : ComputedState dict from state_service.get_computed_state().
    org   : Organisation dict.

    Returns
    -------
    dict with all fields required by every export format.
    """
    from calculations.utilities import kg_to_tonne
    from calculations.gri_framework import run_gap_analysis, gri_coverage_pct, gri_coverage_by_pillar
    from calculations.forecasting import predict_next_emission, annual_projection, detect_trend
    from calculations.benchmarking import benchmark_gap, get_benchmark

    carbon = state.get("carbon", {})
    esg    = state.get("esg",    {})
    dq     = state.get("data_quality", {})
    conf   = state.get("confidence", {})

    org_id  = org.get("org_id", "")
    sector  = org.get("sector", "Manufacturing")
    area_m2 = float(org.get("area_m2", 0) or 0)

    # Attempt GRI gap analysis (requires disclosure inputs from session)
    try:
        from repository.session_repo import get as _get
        di     = _get("disclosure_inputs") or {}
        df_raw = _get("uploaded_df")
        gri    = run_gap_analysis(di, df_raw)
        gri_pct= gri_coverage_pct(gri)
        gri_bp = gri_coverage_by_pillar(gri)
    except Exception:
        gri, gri_pct, gri_bp = [], 0.0, {"E": 0.0, "S": 0.0, "G": 0.0}

    # Forecast — always use the canonical service-level forecast (with DQ integration)
    try:
        df_raw2 = None
        try:
            from repository.session_repo import get_uploaded_df
            df_raw2 = get_uploaded_df()
        except Exception:
            pass
        # Read DQ status for consistent gate behaviour (same as state_service.get_forecast_validation)
        dq_status = None
        try:
            from repository.session_repo import get as _get_r
            vr = _get_r("validation_result")
            if vr and isinstance(vr, dict):
                dq_status = vr.get("status")
        except Exception:
            pass
        from calculations.forecasting import forecast_with_validation, annual_projection, detect_trend
        fcast_validated = forecast_with_validation(df_raw2, dq_validation_status=dq_status)
        forecast        = fcast_validated.get("forecast", {})
        annual_proj     = annual_projection(df_raw2)
        trend           = detect_trend(df_raw2)
    except Exception:
        fcast_validated, forecast, annual_proj, trend = {}, {}, 0.0, {"direction": "insufficient_data"}

    # Benchmark
    bench = float(carbon.get("benchmark") or get_benchmark(sector))
    gap   = carbon.get("gap") or (
        benchmark_gap(carbon.get("intens_m2", 0), bench) if bench > 0 else {}
    )

    scope1_t = round(kg_to_tonne(carbon.get("scope1_kg", 0)), 2)
    scope2_t = round(kg_to_tonne(carbon.get("scope2_kg", 0)), 2)
    scope3_t = round(kg_to_tonne(carbon.get("scope3_kg", 0)), 2)
    total_t  = round(scope1_t + scope2_t + scope3_t, 2)

    return {
        # Identity
        "org_id":           org_id,
        "company":          org.get("company_name", ""),
        "sector":           sector,
        "area_m2":          area_m2,
        "employees":        org.get("employees", 0),
        "province":         org.get("province", ""),
        "reporting_period": org.get("reporting_period", ""),
        "generated_at":     datetime.datetime.now().isoformat(timespec="seconds"),
        "platform_version": "CarbonLens V8",
        # ESG
        "esg":              esg,
        "esg_score":        esg.get("score", 0.0),
        "esg_grade":        esg.get("grade", "--"),
        "esg_label":        esg.get("label", "--"),
        "esg_env":          esg.get("env", 0.0),
        "esg_social":       esg.get("social", 0.0),
        "esg_gov":          esg.get("gov", 0.0),
        "esg_confidence":   esg.get("confidence_score", 0.0),
        "is_provisional":   esg.get("is_provisional", True),
        "methodology_disclaimer": esg.get("methodology_disclaimer", ""),
        # Carbon
        "carbon":           carbon,
        "scope1_tco2e":     scope1_t,
        "scope2_tco2e":     scope2_t,
        "scope3_tco2e":     scope3_t,
        "total_tco2e":      total_t,
        "intensity_kg_m2":  carbon.get("intens_m2", 0.0),
        "scope_source":     carbon.get("scope_source", "none"),
        "pln_ef_used":      carbon.get("pln_ef_used", 0.0),
        # Benchmark
        "benchmark":        bench,
        "gap":              gap,
        "gap_pct":          gap.get("gap_pct", 0.0),
        "above_benchmark":  gap.get("above_benchmark", False),
        # Data Quality
        "dq":               dq,
        "dq_confidence":    dq.get("confidence_score", 0.0),
        "dq_validation":    dq.get("validation_status", "--"),
        "dq_flags":         dq.get("flagged_fields", []),
        # Confidence
        "confidence":       conf,
        # GRI
        "gri":              gri,
        "gri_pct":          gri_pct,
        "gri_by_pillar":    gri_bp,
        # Forecast
        "forecast":         forecast,
        "forecast_valid":   fcast_validated.get("valid", False),
        "forecast_limitation": fcast_validated.get("limitation", ""),
        "annual_tco2e":     round(kg_to_tonne(annual_proj), 2),
        "trend":            trend,
        # ComputedState metadata
        "state_id":         state.get("state_id", ""),
        "state_version":    state.get("version", 0),
        "computed_at":      state.get("computed_at", ""),
    }


def build_snapshot(state: dict, org: dict) -> dict:
    """
    Build a lightweight ReportSnapshot dict for the audit trail.
    Used as the detail payload for report_exported / pdf_generated events.
    Read-only. No side effects.

    Parameters
    ----------
    state : ComputedState dict.
    org   : Organisation dict.

    Returns
    -------
    dict : Compact snapshot capturing key platform state at export time.
    """
    esg    = state.get("esg",          {})
    carbon = state.get("carbon",       {})
    dq     = state.get("data_quality", {})
    from calculations.utilities import kg_to_tonne

    return {
        "snapshot_ts":       datetime.datetime.now().isoformat(timespec="seconds"),
        "state_id":          state.get("state_id", ""),
        "state_version":     state.get("version", 0),
        "company_name":      org.get("company_name", ""),
        "sector":            org.get("sector", ""),
        "reporting_period":  org.get("reporting_period", ""),
        "esg_score":         esg.get("score", 0.0),
        "esg_grade":         esg.get("grade", "--"),
        "is_provisional":    esg.get("is_provisional", True),
        "total_tco2e":       round(kg_to_tonne(carbon.get("total_kg", 0)), 2),
        "intensity_kg_m2":   carbon.get("intens_m2", 0.0),
        "dq_confidence":     dq.get("confidence_score", 0.0),
        "dq_validation":     dq.get("validation_status", "--"),
        "platform_version":  "CarbonLens V8",
    }


def build_csv(context: dict) -> str:
    """
    Build a summary CSV string from the report context.
    Returns a UTF-8 CSV string suitable for st.download_button.
    """
    import csv, io
    buf = io.StringIO()
    w   = csv.writer(buf)

    # Header block
    w.writerow(["CarbonLens V8 — ESG & Carbon Summary Report"])
    w.writerow(["Generated at", context.get("generated_at", "")])
    w.writerow(["Organisation",  context.get("company", "")])
    w.writerow(["Reporting period", context.get("reporting_period", "")])
    w.writerow(["Sector",        context.get("sector", "")])
    w.writerow([])

    # ESG section
    w.writerow(["ESG RESULTS"])
    w.writerow(["Metric", "Value"])
    w.writerow(["ESG Score",           f"{context.get('esg_score', 0):.1f} / 100"])
    w.writerow(["ESG Grade",           context.get("esg_grade", "--")])
    w.writerow(["ESG Status",          "Provisional" if context.get("is_provisional") else "Substantive"])
    w.writerow(["Environmental Score", f"{context.get('esg_env', 0):.1f}"])
    w.writerow(["Social Score",        f"{context.get('esg_social', 0):.1f}"])
    w.writerow(["Governance Score",    f"{context.get('esg_gov', 0):.1f}"])
    w.writerow(["Score Confidence",    f"{context.get('esg_confidence', 0):.0f}%"])
    w.writerow([])

    # Carbon section
    w.writerow(["CARBON ACCOUNTING"])
    w.writerow(["Metric", "Value", "Unit"])
    w.writerow(["Scope 1 (Direct)",    f"{context.get('scope1_tco2e', 0):.2f}", "tCO2e"])
    w.writerow(["Scope 2 (Grid)",      f"{context.get('scope2_tco2e', 0):.2f}", "tCO2e"])
    w.writerow(["Scope 3 (Value chain)",f"{context.get('scope3_tco2e', 0):.2f}","tCO2e"])
    w.writerow(["Total Emissions",     f"{context.get('total_tco2e', 0):.2f}",  "tCO2e"])
    w.writerow(["Carbon Intensity",    f"{context.get('intensity_kg_m2', 0):.2f}", "kg CO2e/m²"])
    w.writerow(["Sector Benchmark",    f"{context.get('benchmark', 0):.0f}",     "kg CO2e/m²"])
    w.writerow(["Gap to Benchmark",    f"{context.get('gap_pct', 0):+.1f}%",    ""])
    w.writerow([])

    # Data Quality
    w.writerow(["DATA QUALITY"])
    w.writerow(["DQ Confidence",       f"{context.get('dq_confidence', 0):.0f}%"])
    w.writerow(["Validation Status",   context.get("dq_validation", "--")])
    w.writerow(["GRI Coverage",        f"{context.get('gri_pct', 0):.1f}%"])

    return buf.getvalue()


def build_json(context: dict) -> str:
    """Build a JSON summary from the report context. Uses NumpyEncoder for safety."""
    from calculations.utilities import NumpyEncoder
    # Remove DataFrame object (not JSON-serialisable)
    safe_ctx = {k: v for k, v in context.items()
                if not hasattr(v, "to_csv")}
    return json.dumps(safe_ctx, cls=NumpyEncoder, indent=2, ensure_ascii=False)


def build_excel(context: dict) -> bytes:
    """
    Build an Excel workbook from the report context.
    Returns bytes suitable for st.download_button(data=..., mime='application/vnd.openxmlformats...').
    Three sheets: Summary, Carbon Detail, ESG Detail.
    """
    import io
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        log.error("openpyxl not installed — Excel export unavailable")
        return b""

    wb    = openpyxl.Workbook()
    _TEAL = "FF0891B2"
    _HEAD = Font(bold=True, color="FFFFFFFF")
    _FILL = PatternFill("solid", fgColor=_TEAL)

    def _header_row(ws, row_idx, values):
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.font  = _HEAD
            cell.fill  = _FILL
            cell.alignment = Alignment(horizontal="center")

    # ── Sheet 1: Summary ─────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Summary"
    ws1.append(["CarbonLens V8 ESG & Carbon Report"])
    ws1.append(["Organisation", context.get("company", "")])
    ws1.append(["Reporting period", context.get("reporting_period", "")])
    ws1.append(["Generated", context.get("generated_at", "")])
    ws1.append([])
    _header_row(ws1, 6, ["Metric", "Value"])
    summary_rows = [
        ("ESG Score",          f"{context.get('esg_score', 0):.1f} / 100"),
        ("ESG Grade",          context.get("esg_grade", "--")),
        ("ESG Status",         "Provisional" if context.get("is_provisional") else "Substantive"),
        ("Total Emissions",    f"{context.get('total_tco2e', 0):.2f} tCO2e"),
        ("Carbon Intensity",   f"{context.get('intensity_kg_m2', 0):.2f} kg/m²"),
        ("Benchmark Gap",      f"{context.get('gap_pct', 0):+.1f}%"),
        ("DQ Confidence",      f"{context.get('dq_confidence', 0):.0f}%"),
        ("GRI Coverage",       f"{context.get('gri_pct', 0):.1f}%"),
    ]
    for row in summary_rows:
        ws1.append(row)
    ws1.column_dimensions["A"].width = 25
    ws1.column_dimensions["B"].width = 25

    # ── Sheet 2: Carbon Detail ────────────────────────────────────────────────
    ws2 = wb.create_sheet("Carbon")
    _header_row(ws2, 1, ["Scope", "Emissions (tCO2e)", "% of Total"])
    total = context.get("total_tco2e", 0) or 1
    for scope, val in [
        ("Scope 1 — Direct Combustion", context.get("scope1_tco2e", 0)),
        ("Scope 2 — Grid Electricity",  context.get("scope2_tco2e", 0)),
        ("Scope 3 — Value Chain",       context.get("scope3_tco2e", 0)),
    ]:
        ws2.append([scope, round(val, 2), f"{val/total*100:.1f}%"])
    ws2.append(["Total", context.get("total_tco2e", 0), "100%"])

    # ── Sheet 3: ESG Detail ───────────────────────────────────────────────────
    ws3 = wb.create_sheet("ESG")
    _header_row(ws3, 1, ["Pillar", "Score", "Weight"])
    for pillar, score, wt in [
        ("Environmental", context.get("esg_env",    0), "40%"),
        ("Social",        context.get("esg_social", 0), "30%"),
        ("Governance",    context.get("esg_gov",    0), "30%"),
        ("Composite",     context.get("esg_score",  0), "100%"),
    ]:
        ws3.append([pillar, round(score, 1), wt])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
